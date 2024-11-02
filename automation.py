import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

def process_workbook(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb ['Sheet1']

    for row in range(2, sheet.max_row + 1):
        goals = sheet.cell(row, 3)
        assists = sheet.cell(row, 4)
        total_contributions = goals.value + assists.value
        total_contributions_cell = sheet.cell(row, 5)
        total_contributions_cell.value = total_contributions

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=5,
              max_col=5)

    chart = BarChart()
    chart.add_data(values)
    chart.title = 'Contributions'
    sheet.add_chart(chart, 'g2')

    #Make the Header Row bold and change its color
    font_style = Font(bold=True, color='ff5334')

    for cell in sheet[1]:
        cell.font = font_style

    wb.save(file_name)


process_workbook('Football.xlsx')

